import string
from openpyxl import load_workbook

class ExcelDeal(object):
    def __init__(self,file_path,sheet_name=None):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def read_col(self,start_col = 0,end_col="",choose_col=[]):
        column_lists = list(self.sheet.columns)
        choose_col_lists = [column_lists[col-1] for col in choose_col]

        col_num = 0
        while True:
            if start_col > col_num:
                continue
            every_row = [str(col[col_num].value).strip() for col in choose_col_lists]

    def read_row(self,start_row = 1,choose_attr = [],show_stats="str"):
        """
        :param start_row:
        :param choose_col:
        :param show_stats:
            str   .value.strip()
            ori   .value
            None  col[]
        :return:
        """
        col_indexs = []
        attr_sorted = None
        if choose_attr:
            if type(choose_attr) == dict:
                choose_col = choose_attr.keys()
                attr_sorted = [value for key,value in sorted(choose_attr.items(), key=lambda ite: ite[0])]
            elif type(choose_attr) == list:
                choose_col = choose_attr
            else:
                choose_col = []

            upper_letter = list(string.ascii_uppercase)
            if all([str(i).isdigit() and type(i) == int for i in choose_col]):
                col_indexs = sorted(set(choose_col))
            elif all([str(j).isupper() and j in upper_letter for j in choose_col]):
                col_indexs = [upper_letter.index(col_letter.upper())+1 for col_letter in choose_col]
                col_indexs = sorted(set(col_indexs))
            elif len(choose_col) == 0:
                pass
            else:
                assert 0,f"{choose_col} is trouble !"

        for index, col in enumerate(self.read().rows):
            if index < start_row - 1:
                continue
            if col_indexs:
                col = [col[col_index-1] for col_index in col_indexs]

            if show_stats == "str":
                col = [str(every_grid.value if every_grid.value != None else "").strip() for every_grid in col]
            elif show_stats == "ori":
                col = [every_grid.value for every_grid in col]

            if attr_sorted:
                dic = {}
                for index,value in enumerate(col):
                    dic[attr_sorted[index]] = value
                col = dic

            yield col

    def read(self):
        sheet_name = self.sheet_name
        workbook_ = load_workbook(filename=self.file_path,data_only=True)
        sheetnames = workbook_.get_sheet_names()
        # sheetname = ""
        if not sheet_name:
            sheetname = sheetnames[0]
        else:
            sheetname = sheet_name if sheet_name in sheetnames else ""
        assert sheetname,"Please choose right sheetname !"
        sheet = workbook_[sheetname]

        return sheet